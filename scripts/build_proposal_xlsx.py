#!/usr/bin/env python3
"""Client-facing project proposal workbook (no code) for B2B_Marketplace."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

TEAL="1B6E85"; TEAL_DARK="0C3B49"; CORAL="F0875A"; GREEN="1E8E5A"
LIGHT="EAF2F4"; ZEBRA="F6F9FA"; WHITE="FFFFFF"; GREY="5B6B73"
COREBG="E4F4EC"; CORE="1E8E5A"; P2BG="FBF3E2"; P2="B7791F"
GANTT="1B6E85"; GANTT2="3FA9C0"

thin=Side(style="thin",color="D6DEE2")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
def fill(h): return PatternFill("solid",fgColor=h)

def title_block(ws,title,subtitle,span):
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=span)
    c=ws.cell(1,1,title); c.font=Font(size=18,bold=True,color=WHITE); c.fill=fill(TEAL)
    c.alignment=Alignment(vertical="center",indent=1); ws.row_dimensions[1].height=38
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=span)
    s=ws.cell(2,1,subtitle); s.font=Font(size=10,italic=True,color=WHITE); s.fill=fill(TEAL_DARK)
    s.alignment=Alignment(vertical="center",indent=1); ws.row_dimensions[2].height=20

def header_row(ws,row,headers):
    for i,h in enumerate(headers,1):
        c=ws.cell(row,i,h); c.font=Font(bold=True,color=WHITE,size=11); c.fill=fill(TEAL)
        c.alignment=Alignment(vertical="center",indent=1,wrap_text=True); c.border=border
    ws.row_dimensions[row].height=26

def cell(ws,r,c,v,bold=False,color="333333",zebra=None,center=False,size=10,wrap=True):
    x=ws.cell(r,c,v)
    x.font=Font(bold=bold,color=color,size=size)
    x.alignment=Alignment(vertical="center",horizontal="center" if center else "left",
                          indent=0 if center else 1,wrap_text=wrap)
    x.border=border
    if zebra: x.fill=fill(zebra)
    return x

def priority_cell(ws,r,c,val):
    x=ws.cell(r,c,val); x.alignment=Alignment(horizontal="center",vertical="center"); x.border=border
    if val.lower().startswith("core"):
        x.fill=fill("E4F4EC"); x.font=Font(color="1E8E5A",bold=True,size=10)
    else:
        x.fill=fill("FBF3E2"); x.font=Font(color="B7791F",bold=True,size=10)

wb=Workbook()

# ====================================================== 1. OVERVIEW
ov=wb.active; ov.title="1. Project Overview"; ov.sheet_view.showGridLines=False
title_block(ov,"B2B Marketplace Platform","Project Proposal & Technical Approach  ·  Prepared for Client",2)
info=[
 ("Project","B2B Marketplace (verified supplier directory + RFQ / quotation engine)"),
 ("Objective","Connect verified suppliers with buyers; enable RFQs, quotations and a credit-based contact-unlock model"),
 ("Primary users","Buyer  ·  Supplier  ·  Administrator"),
 ("Approach","Modern, SEO-first, all-TypeScript web platform (Next.js + PostgreSQL)"),
 ("Delivery model","Phased delivery in 8 milestones with review at each stage"),
 ("",""),
 ("Scope — Public","Marketing site, supplier directory, product catalog, categories, search & filters, SEO"),
 ("Scope — Buyer","RFQs, quotations, credits, saved items, notifications, profile"),
 ("Scope — Supplier","RFQ feed, quotations, product catalog, company profile, verification"),
 ("Scope — Admin","User & supplier management, moderation, CMS, FAQ, categories, credits, audit"),
 ("",""),
 ("Estimated duration","~14 weeks (≈ 3.5 months), excluding client review cycles"),
 ("Feature modules","8 functional modules · 59 features"),
 ("Note","This document is a planning & technical-approach proposal. No source code is included."),
]
r=4
for k,v in info:
    a=ov.cell(r,1,k); b=ov.cell(r,2,v)
    a.font=Font(bold=True,color=TEAL_DARK,size=11); a.fill=fill(LIGHT)
    a.alignment=Alignment(vertical="center",indent=1)
    b.font=Font(size=11,color="222222"); b.alignment=Alignment(vertical="center",indent=1,wrap_text=True)
    a.border=border; b.border=border; ov.row_dimensions[r].height=24; r+=1
ov.column_dimensions["A"].width=22; ov.column_dimensions["B"].width=92

# ====================================================== 2. FEATURE LIST
fe=wb.create_sheet("2. Feature List"); fe.sheet_view.showGridLines=False
title_block(fe,"Feature List","Complete scope grouped by area",4)
header_row(fe,4,["#","Area","Feature","Priority"]); fe.freeze_panes="A5"
CORE="Core (MVP)"; P2L="Phase 2"
feats=[
 ("Public Website","Home / landing page with hero & search",CORE),
 ("Public Website","About, Contact & FAQ pages",CORE),
 ("Public Website","Supplier directory with filters",CORE),
 ("Public Website","Supplier profile / details page",CORE),
 ("Public Website","Product catalog with advanced filters",CORE),
 ("Public Website","Product details page",CORE),
 ("Public Website","Category system (parent / child / dynamic)",CORE),
 ("Public Website","Filters: country, MOQ, budget, verified",CORE),
 ("Public Website","Country-based pricing & currency conversion",CORE),
 ("Public Website","SEO (metadata, structured data, sitemap)",CORE),
 ("Public Website","Legal / content (Privacy, Terms, Cookie)",CORE),
 ("Public Website","Responsive design + dark mode",CORE),
 ("Authentication","Email / password login",CORE),
 ("Authentication","Buyer registration",CORE),
 ("Authentication","Supplier registration (multi-step)",CORE),
 ("Authentication","Google single sign-on",CORE),
 ("Authentication","Forgot / reset password",CORE),
 ("Authentication","Phone OTP verification",P2L),
 ("Authentication","Roles & protected areas",CORE),
 ("Buyer","Dashboard overview",CORE),
 ("Buyer","Post & manage RFQs",CORE),
 ("Buyer","Receive & accept / reject quotations",CORE),
 ("Buyer","Unlock supplier contacts (credits)",CORE),
 ("Buyer","Buy credits & view transaction history",CORE),
 ("Buyer","Saved suppliers / products / RFQs",CORE),
 ("Buyer","Notifications & preferences",CORE),
 ("Buyer","Profile settings & change password",CORE),
 ("Buyer","Switch between buyer / supplier mode",P2L),
 ("Supplier","Dashboard overview",CORE),
 ("Supplier","Browse RFQ feed",CORE),
 ("Supplier","Submit & withdraw quotations",CORE),
 ("Supplier","Product management (CRUD)",CORE),
 ("Supplier","Company profile & trade terms",CORE),
 ("Supplier","Verification (document submission)",CORE),
 ("Supplier","Profile settings & change password",CORE),
 ("Admin","Dashboard with stats & activity",CORE),
 ("Admin","User management (approve / block)",CORE),
 ("Admin","Impersonate user login",CORE),
 ("Admin","Supplier verification queue",CORE),
 ("Admin","Document review (approve / reject)",CORE),
 ("Admin","RFQ moderation (flag / delete)",CORE),
 ("Admin","Category management",CORE),
 ("Admin","CMS page management",CORE),
 ("Admin","FAQ management",CORE),
 ("Admin","Credit / token management",CORE),
 ("Admin","Audit logging",CORE),
 ("Admin","Roles & permissions (granular RBAC)",P2L),
 ("Payments & Credits","Credit packs",CORE),
 ("Payments & Credits","Stripe checkout",CORE),
 ("Payments & Credits","Payment webhook & credit grant",CORE),
 ("Payments & Credits","Credit ledger (transaction-safe)",CORE),
 ("Notifications","In-app + notification center",CORE),
 ("Notifications","Email notifications",CORE),
 ("Notifications","SMS / WhatsApp (Twilio)",P2L),
 ("Notifications","Telegram notifications",P2L),
 ("Notifications","Per-channel preferences",CORE),
 ("Platform","Multi-currency support",CORE),
 ("Platform","Security & data integrity",CORE),
]
r=5
for i,(area,feat,prio) in enumerate(feats,1):
    z=ZEBRA if i%2==0 else WHITE
    cell(fe,r,1,i,center=True,zebra=z)
    cell(fe,r,2,area,bold=True,color=TEAL_DARK,zebra=z)
    cell(fe,r,3,feat,zebra=z)
    priority_cell(fe,r,4,prio)
    fe.row_dimensions[r].height=22; r+=1
for col,w in zip("ABCD",[5,22,60,16]): fe.column_dimensions[col].width=w
lr=r+1
fe.cell(lr,2,"Legend:").font=Font(bold=True,color=GREY)
priority_cell(fe,lr,3,"Core (MVP)"); priority_cell(fe,lr,4,"Phase 2")

# ====================================================== 3. MODULES & FUNCTIONALITIES
mo=wb.create_sheet("3. Modules & Functionalities"); mo.sheet_view.showGridLines=False
title_block(mo,"Modules & Functionalities","System decomposed into modules and the roles they serve",4)
header_row(mo,4,["Module","Functionality","Role(s)","Description"]); mo.freeze_panes="A5"
mods=[
 ("Catalog & Discovery","Supplier directory","Public","Searchable, filterable list of verified suppliers"),
 ("Catalog & Discovery","Product catalog","Public","Browse products with localized pricing & filters"),
 ("Catalog & Discovery","Category system","Public / Admin","Parent/child categories with dynamic pages"),
 ("RFQ & Quotation","Post RFQ","Buyer","Create requests for quotation"),
 ("RFQ & Quotation","RFQ feed","Supplier","Discover open requests to quote on"),
 ("RFQ & Quotation","Quotations","Buyer / Supplier","Submit, compare, accept or reject quotes"),
 ("Monetization","Credit wallet","Buyer","Buy credits; view transaction history"),
 ("Monetization","Unlock contacts","Buyer","Spend credits to reveal supplier contacts"),
 ("Monetization","Payments","Buyer","Stripe checkout + webhook credit grant"),
 ("Supplier Management","Product management","Supplier","Create / edit / delete catalog products"),
 ("Supplier Management","Company profile","Supplier","Company info, terms, strengths, contacts"),
 ("Supplier Management","Verification","Supplier / Admin","Submit documents; admin approves badge"),
 ("Identity & Access","Authentication","All","Login, registration, SSO, password reset"),
 ("Identity & Access","Roles & permissions","All","Buyer / Supplier / Admin access control"),
 ("Engagement","Saved items","Buyer","Bookmark suppliers, products, RFQs"),
 ("Engagement","Notifications","All","In-app + email / SMS / WhatsApp / Telegram"),
 ("Administration","User management","Admin","Approve, block, impersonate users"),
 ("Administration","Moderation","Admin","Flag / delete spam RFQs"),
 ("Administration","Content (CMS & FAQ)","Admin","Manage pages and FAQs"),
 ("Administration","Credit management","Admin","Adjust balances; review transactions"),
 ("Administration","Audit log","Admin","Record of all administrative actions"),
 ("Platform Services","SEO engine","Public","Metadata, structured data, sitemaps"),
 ("Platform Services","Currency & geo","Public","Country detection & price conversion"),
]
r=5
for i,(m,f,role,desc) in enumerate(mods,1):
    z=ZEBRA if i%2==0 else WHITE
    cell(mo,r,1,m,bold=True,color=TEAL_DARK,zebra=z)
    cell(mo,r,2,f,bold=True,zebra=z,color="222222")
    cell(mo,r,3,role,zebra=z,center=False)
    cell(mo,r,4,desc,zebra=z)
    mo.row_dimensions[r].height=24; r+=1
for col,w in zip("ABCD",[26,26,22,58]): mo.column_dimensions[col].width=w

# ====================================================== 4. TECHNOLOGY STACK
te=wb.create_sheet("4. Technology Stack"); te.sheet_view.showGridLines=False
title_block(te,"Technology Stack & Purpose","Tools proposed, why they were chosen, and where they're used",4)
header_row(te,4,["Area","Technology","Purpose / Why chosen","Used for"]); te.freeze_panes="A5"
tech=[
 ("Framework","Next.js (App Router) + React","SEO-friendly server rendering; single full-stack codebase","Entire web app"),
 ("Language","TypeScript","Type safety reduces bugs across the stack","Whole project"),
 ("Backend","Next.js Server Actions / API routes","Node.js backend without a separate API server","Business logic & APIs"),
 ("Database","PostgreSQL","Relational integrity for RFQs, quotes, credits","All persistent data"),
 ("ORM","Prisma","Type-safe data access, schema & migrations","Database layer"),
 ("Authentication","Auth.js (NextAuth)","Email, Google SSO, roles, impersonation","Login & access control"),
 ("Validation","Zod","One schema validates client + server input","Forms & APIs"),
 ("UI / Styling","Tailwind CSS + shadcn/ui","Custom, accessible, modern interface","All screens"),
 ("Animation","GSAP","Smooth scroll-reveal & hero animation","Marketing pages"),
 ("3D Graphics","Three.js / React Three Fiber","Interactive 3D hero for premium feel","Home page"),
 ("Smooth Scroll","Lenis","Refined scrolling experience","Public site"),
 ("Payments","Stripe","Secure card checkout for credits","Credit purchase"),
 ("Email","Resend","Transactional email","Notifications"),
 ("SMS / WhatsApp","Twilio","Phone verification & alerts","Notifications & OTP"),
 ("Messaging","Telegram Bot API","Opt-in Telegram alerts","Notifications"),
 ("File Storage","Cloudflare R2 / S3","Documents & media storage","Verification & images"),
 ("Currency / Geo","Exchange-rate API + IP geo","Localized pricing & conversion","Pricing"),
 ("Hosting","Any Node.js host","Deploys to Railway / Render / DO / AWS / VPS","Production hosting"),
 ("Database Hosting","Managed PostgreSQL","Scalable managed database","Production DB"),
 ("Monitoring","Sentry + Analytics","Error tracking & performance","Operations"),
]
r=5
for i,(a,t,why,used) in enumerate(tech,1):
    z=ZEBRA if i%2==0 else WHITE
    cell(te,r,1,a,bold=True,color=TEAL_DARK,zebra=z)
    cell(te,r,2,t,bold=True,zebra=z,color="222222")
    cell(te,r,3,why,zebra=z)
    cell(te,r,4,used,zebra=z)
    te.row_dimensions[r].height=24; r+=1
for col,w in zip("ABCD",[22,34,56,28]): te.column_dimensions[col].width=w

# ====================================================== 5. SYSTEM ARCHITECTURE
ar=wb.create_sheet("5. System Architecture"); ar.sheet_view.showGridLines=False
title_block(ar,"System Architecture Overview","Layered design, components and data flow",2)
# Diagram (text)
ar.merge_cells("A4:B4")
d=ar.cell(4,1,"Architecture (high level)"); d.font=Font(bold=True,color=TEAL,size=12)
diagram=[
 "┌──────────────────────────────────────────────────────────┐",
 "│  CLIENT (Browser / Mobile web)                            │",
 "│  Public site (SEO)  ·  Buyer & Supplier dashboards  ·  Admin │",
 "└───────────────────────────┬──────────────────────────────┘",
 "                            │  HTTPS",
 "┌───────────────────────────▼──────────────────────────────┐",
 "│  APPLICATION  (Next.js — server rendering + API layer)    │",
 "│  Auth & access control · Validation · Business logic      │",
 "└──────┬───────────────────────────────────┬───────────────┘",
 "       │ Prisma ORM                         │ Integrations",
 "┌──────▼───────┐                  ┌──────────▼──────────────┐",
 "│ PostgreSQL   │                  │ Stripe · Resend · Twilio │",
 "│  Database    │                  │ Storage · Currency API   │",
 "└──────────────┘                  └─────────────────────────┘",
]
r=5
for line in diagram:
    ar.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    c=ar.cell(r,1,line); c.font=Font(name="Courier New",size=9,color=TEAL_DARK)
    c.alignment=Alignment(vertical="center"); ar.row_dimensions[r].height=15; r+=1
# Components table
r+=1
ar.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
h=ar.cell(r,1,"Layer responsibilities"); h.font=Font(bold=True,color=TEAL,size=12); r+=1
header_row(ar,r,["Layer","Responsibility"]); r+=1
layers=[
 ("Presentation","Server-rendered, SEO-optimized pages and role-based dashboards (Buyer/Supplier/Admin)"),
 ("Application / API","Request handling, authentication, input validation and all business logic"),
 ("Data access","Type-safe database access and migrations via the ORM"),
 ("Database","PostgreSQL — relational storage for users, products, RFQs, quotes, credits"),
 ("Integrations","Payments (Stripe), email/SMS (Resend/Twilio), storage (R2/S3), currency API"),
 ("Security","Role-based access, encrypted passwords, transaction-safe operations, audit log"),
]
for i,(l,resp) in enumerate(layers):
    z=ZEBRA if i%2 else WHITE
    cell(ar,r,1,l,bold=True,color=TEAL_DARK,zebra=z); cell(ar,r,2,resp,zebra=z)
    ar.row_dimensions[r].height=26; r+=1
# Flow
r+=1
ar.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
h=ar.cell(r,1,"Core flow — RFQ to deal"); h.font=Font(bold=True,color=TEAL,size=12); r+=1
for step in [
 "1.  Buyer posts an RFQ  →  stored and broadcast to the supplier feed",
 "2.  Suppliers submit quotations  →  buyer is notified",
 "3.  Buyer accepts a quote  →  others auto-declined, supplier notified",
 "4.  Buyer spends credits to unlock supplier contact  →  deal proceeds off-platform",
]:
    ar.merge_cells(start_row=r,start_column=1,end_row=r,end_column=2)
    c=ar.cell(r,1,step); c.font=Font(size=10,color="333333"); c.alignment=Alignment(indent=1,vertical="center")
    ar.row_dimensions[r].height=18; r+=1
ar.column_dimensions["A"].width=26; ar.column_dimensions["B"].width=88

# ====================================================== 6. DEVELOPMENT PHASES
dp=wb.create_sheet("6. Development Phases"); dp.sheet_view.showGridLines=False
title_block(dp,"Development Phases","Phased delivery with clear milestones",4)
header_row(dp,4,["Phase","Milestone","Key Deliverables","Duration"]); dp.freeze_panes="A5"
phases=[
 ("0","Foundation","Project setup, database design, design system, auth base","1 week"),
 ("1","Public SEO Website","Listings, detail pages, categories, filters, currency, SEO","2 weeks"),
 ("2","Authentication & Onboarding","Login, registration, Google SSO, supplier wizard","2 weeks"),
 ("3","Buyer Dashboard","RFQs, quotations, saved items, credits, profile","2 weeks"),
 ("4","Supplier Dashboard","RFQ feed, quotes, products, company, verification","2 weeks"),
 ("5","Payments & Credits","Unlock flow, Stripe checkout, credit packs","1 week"),
 ("6","Admin Panel","Users, verification, CMS, FAQ, categories, credits","2 weeks"),
 ("7","Notifications, QA & Deployment","Notification center, testing, deployment, handover","2 weeks"),
]
r=5
for i,(ph,ms,deliv,dur) in enumerate(phases):
    z=ZEBRA if i%2 else WHITE
    cell(dp,r,1,ph,center=True,bold=True,color=TEAL_DARK,zebra=z)
    cell(dp,r,2,ms,bold=True,zebra=z,color="222222")
    cell(dp,r,3,deliv,zebra=z)
    cell(dp,r,4,dur,center=True,zebra=z,color=TEAL_DARK,bold=True)
    dp.row_dimensions[r].height=26; r+=1
cell(dp,r,2,"Total estimated duration",bold=True,color=TEAL_DARK)
cell(dp,r,3,"Excludes client review / UAT cycles")
cell(dp,r,4,"~14 weeks",center=True,bold=True,color=GREEN)
for col,w in zip("ABCD",[8,30,72,14]): dp.column_dimensions[col].width=w

# ====================================================== 7. TIMELINE (GANTT)
tl=wb.create_sheet("7. Estimated Timeline"); tl.sheet_view.showGridLines=False
WEEKS=14
title_block(tl,"Estimated Timeline","Indicative schedule — week-by-week (Gantt)",2+WEEKS)
# header
hdr=["Phase","Weeks"]+[f"W{i}" for i in range(1,WEEKS+1)]
header_row(tl,4,hdr); tl.freeze_panes="C5"
gantt=[
 ("0 — Foundation",1,1),
 ("1 — Public Site",2,3),
 ("2 — Authentication",4,5),
 ("3 — Buyer Dashboard",6,7),
 ("4 — Supplier Dashboard",8,9),
 ("5 — Payments & Credits",10,10),
 ("6 — Admin Panel",11,12),
 ("7 — Notifications, QA & Deploy",13,14),
]
r=5
for i,(name,ws_,we) in enumerate(gantt):
    z=ZEBRA if i%2 else WHITE
    cell(tl,r,1,name,bold=True,color=TEAL_DARK,zebra=z)
    cell(tl,r,2,f"W{ws_}–W{we}" if we>ws_ else f"W{ws_}",center=True,zebra=z,color=GREY,size=9)
    for w in range(1,WEEKS+1):
        x=tl.cell(r,2+w); x.border=border
        if ws_<=w<=we:
            x.fill=fill(GANTT if i%2==0 else GANTT2)
        else:
            x.fill=fill(z)
    tl.row_dimensions[r].height=22; r+=1
# month guide
r+=1
cell(tl,r,1,"Approx. months",bold=True,color=TEAL_DARK)
spans=[("Month 1",1,4),("Month 2",5,8),("Month 3",9,12),("Month 3.5",13,14)]
for label,a,b in spans:
    tl.merge_cells(start_row=r,start_column=2+a,end_row=r,end_column=2+b)
    c=tl.cell(r,2+a,label); c.fill=fill(LIGHT); c.font=Font(bold=True,color=TEAL_DARK,size=9)
    c.alignment=Alignment(horizontal="center",vertical="center"); c.border=border
tl.column_dimensions["A"].width=30; tl.column_dimensions["B"].width=12
for w in range(1,WEEKS+1): tl.column_dimensions[chr(ord("C")+w-1)].width=5

out="B2B_Marketplace_Project_Proposal.xlsx"
wb.save(out)
print("Saved:",out,"| Sheets:",len(wb.sheetnames))
for s in wb.sheetnames: print("  -",s)
