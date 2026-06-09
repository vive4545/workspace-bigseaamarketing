#!/usr/bin/env python3
"""Plain, unstyled project-plan workbook scoped to Phases 2-6.
No per-phase dates. Single final delivery date stated in the Overview."""
from openpyxl import Workbook
from openpyxl.styles import Font

bold = Font(bold=True)

FINAL_DELIVERY = "26 June 2026"

# Phase: number, name, deliverables
phase_defs = [
    (2, "Authentication & Onboarding",
     "Login, registration, Google SSO, supplier onboarding, password reset"),
    (3, "Buyer Dashboard",
     "RFQs, quotations, saved items, notifications, profile"),
    (4, "Supplier Dashboard",
     "RFQ feed, quotations, product management, company profile, verification"),
    (5, "Payments & Credits",
     "Credit packs, Stripe checkout, payment webhook, credit ledger, unlock contacts"),
    (6, "Admin Panel",
     "User management, verification, moderation, CMS, FAQ, categories, credits, audit"),
]
phase_name = {n: nm for n, nm, _ in phase_defs}

features = {
    2: ["Login (email & password)", "Buyer registration", "Supplier registration (multi-step)",
        "Google single sign-on", "Forgot / reset password", "Phone OTP verification",
        "User roles & protected areas"],
    3: ["Dashboard overview", "Post RFQ", "Manage RFQs (edit / close / delete)",
        "Receive & accept / reject quotations", "Saved suppliers / products / RFQs",
        "Notifications & preferences", "Profile settings & change password"],
    4: ["Dashboard overview", "Browse RFQ feed", "Submit & withdraw quotations",
        "Product management (add / edit / delete)", "Company profile & trade terms",
        "Verification (document submission)", "Profile settings & change password"],
    5: ["Credit packs", "Buy credits (Stripe checkout)", "Payment webhook & credit grant",
        "Credit ledger / transaction history", "Unlock supplier contacts"],
    6: ["Admin dashboard (stats & activity)", "User management (approve / block)",
        "Impersonate user login", "Supplier verification queue", "Document review (approve / reject)",
        "RFQ moderation (flag / delete)", "Category management", "CMS page management",
        "FAQ management", "Credit / token management", "Audit logging"],
}

technology = [
    ("Framework", "Next.js / React", "SEO-friendly server rendering; full-stack app"),
    ("Language", "TypeScript", "Type safety across the application"),
    ("Database", "PostgreSQL", "Relational data for users, RFQs, quotes, credits"),
    ("ORM", "Prisma", "Type-safe database access & migrations"),
    ("Authentication", "Auth.js (NextAuth)", "Email, Google SSO, roles, impersonation"),
    ("Validation", "Zod", "Client & server input validation"),
    ("UI / Styling", "Tailwind CSS + shadcn/ui", "Modern, accessible interface"),
    ("Payments", "Stripe", "Secure credit-pack checkout"),
    ("Email", "Resend", "Transactional email"),
    ("SMS / WhatsApp", "Twilio", "Phone verification & alerts"),
    ("Messaging", "Telegram Bot API", "Opt-in notifications"),
    ("File Storage", "Cloudflare R2 / S3", "Documents & media"),
    ("Hosting", "Any Node.js host", "Railway / Render / DigitalOcean / AWS / VPS"),
    ("Database Hosting", "Managed PostgreSQL", "Scalable managed database"),
]

wb = Workbook()


def widths(ws, ws_widths):
    for col, w in ws_widths.items():
        ws.column_dimensions[col].width = w


# ===== Sheet 1: Overview =====
ov = wb.active
ov.title = "Overview"
rows = [
    ("Project", "B2B Marketplace Platform"),
    ("Scope", "Phases 2 to 6"),
    ("Phases included", "2 Authentication, 3 Buyer Dashboard, 4 Supplier Dashboard, 5 Payments & Credits, 6 Admin Panel"),
    ("Users", "Buyer, Supplier, Admin"),
    ("Final delivery date", FINAL_DELIVERY),
]
for i, (k, v) in enumerate(rows, 1):
    ov.cell(i, 1, k).font = bold
    ov.cell(i, 2, v)
widths(ov, {"A": 20, "B": 80})

# ===== Sheet 2: Modules & Features =====
mf = wb.create_sheet("Modules & Features")
for c, h in enumerate(["Phase", "Module", "Feature"], 1):
    mf.cell(1, c, h).font = bold
r = 2
for num, *_ in phase_defs:
    for feat in features[num]:
        mf.cell(r, 1, num)
        mf.cell(r, 2, phase_name[num])
        mf.cell(r, 3, feat)
        r += 1
widths(mf, {"A": 8, "B": 28, "C": 55})

# ===== Sheet 3: Technology Stack =====
ts = wb.create_sheet("Technology Stack")
for c, h in enumerate(["Area", "Technology", "Purpose"], 1):
    ts.cell(1, c, h).font = bold
for i, (a, t, p) in enumerate(technology, 2):
    ts.cell(i, 1, a)
    ts.cell(i, 2, t)
    ts.cell(i, 3, p)
widths(ts, {"A": 22, "B": 28, "C": 55})

# ===== Sheet 4: Development Phases =====
dp = wb.create_sheet("Development Phases")
for c, h in enumerate(["Phase", "Module", "Key Deliverables"], 1):
    dp.cell(1, c, h).font = bold
for i, (num, name, deliv) in enumerate(phase_defs, 2):
    dp.cell(i, 1, num)
    dp.cell(i, 2, name)
    dp.cell(i, 3, deliv)
note = len(phase_defs) + 3
dp.cell(note, 1, "Final delivery date").font = bold
dp.cell(note, 3, FINAL_DELIVERY)
widths(dp, {"A": 8, "B": 28, "C": 70})

out = "B2B_Marketplace_Project_Plan.xlsx"
wb.save(out)
print("Saved:", out, "| Sheets:", wb.sheetnames)
print("Final delivery date:", FINAL_DELIVERY)
