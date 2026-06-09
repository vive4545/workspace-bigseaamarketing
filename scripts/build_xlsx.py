#!/usr/bin/env python3
"""Generate a client-ready Excel workbook: BigSeaa feature list + technology."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Brand palette ---------------------------------------------------------
TEAL = "1B6E85"
TEAL_DARK = "0C3B49"
CORAL = "F0875A"
GREEN = "1E8E5A"
GREEN_BG = "E4F4EC"
AMBER = "B7791F"
AMBER_BG = "FBF3E2"
LIGHT = "EAF2F4"
ZEBRA = "F6F9FA"
WHITE = "FFFFFF"
GREY = "5B6B73"

thin = Side(style="thin", color="D6DEE2")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def fill(hex_):
    return PatternFill("solid", fgColor=hex_)


def title_block(ws, title, subtitle, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(1, 1, title)
    c.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
    c.fill = fill(TEAL)
    c.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 38

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)
    s = ws.cell(2, 1, subtitle)
    s.font = Font(name="Calibri", size=10, italic=True, color=WHITE)
    s.fill = fill(TEAL_DARK)
    s.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 20


def header_row(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row, i, h)
        c.font = Font(bold=True, color=WHITE, size=11)
        c.fill = fill(TEAL)
        c.alignment = Alignment(vertical="center", horizontal="left", indent=1, wrap_text=True)
        c.border = border
    ws.row_dimensions[row].height = 26


def status_style(cell, value):
    v = value.lower()
    if "ready" in v or "config" in v or "needs" in v:
        cell.fill = fill(AMBER_BG); cell.font = Font(color=AMBER, bold=True, size=10)
    else:
        cell.fill = fill(GREEN_BG); cell.font = Font(color=GREEN, bold=True, size=10)
    cell.alignment = Alignment(vertical="center", horizontal="center")
    cell.border = border


wb = Workbook()

# ============================================================ OVERVIEW
ov = wb.active
ov.title = "Overview"
ov.sheet_view.showGridLines = False
title_block(ov, "BigSeaa — B2B Marketplace Platform",
            "Feature List & Technology Stack  ·  Prepared for Client", 2)

rows = [
    ("Project", "BigSeaa — B2B Marketplace (supplier directory + RFQ/quotation engine)"),
    ("Personas", "Buyer  ·  Supplier  ·  Admin"),
    ("Architecture", "Next.js 16 full-stack (SSR/SEO) + PostgreSQL + Prisma"),
    ("Frontend", "Next.js / React 19 / TypeScript / Tailwind CSS"),
    ("Backend", "Next.js Route Handlers + Server Actions (Node.js)"),
    ("Database", "PostgreSQL (relational) via Prisma ORM"),
    ("Animation / UX", "GSAP · Three.js (3D hero) · Lenis smooth scroll"),
    ("Payments", "Stripe (credit packs + webhooks)"),
    ("Notifications", "Email · SMS · WhatsApp · Telegram (multi-channel)"),
    ("Hosting", "Any Node.js host (Railway, Render, DigitalOcean, AWS, VPS, etc.)"),
    ("", ""),
    ("Page routes", "42"),
    ("Data models", "28"),
    ("Feature modules", "8 (Public, Auth, Buyer, Supplier, Admin, Payments, Notifications, Platform)"),
    ("Status", "All core features delivered & build-verified"),
]
r = 4
for k, v in rows:
    a = ov.cell(r, 1, k); b = ov.cell(r, 2, v)
    a.font = Font(bold=True, color=TEAL_DARK, size=11)
    a.fill = fill(LIGHT); a.alignment = Alignment(vertical="center", indent=1)
    b.font = Font(size=11, color="222222")
    b.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    a.border = border; b.border = border
    ov.row_dimensions[r].height = 22
    r += 1
ov.column_dimensions["A"].width = 22
ov.column_dimensions["B"].width = 84
ov.sheet_view.showGridLines = False

# ============================================================ FEATURES
fe = wb.create_sheet("Feature List")
fe.sheet_view.showGridLines = False
title_block(fe, "Feature List", "Complete scope by module  ·  Status as delivered", 5)
header_row(fe, 4, ["#", "Module", "Feature", "Description", "Status"])
fe.freeze_panes = "A5"

DONE = "Completed"
READY = "Ready – needs API key"

features = [
    # Public
    ("Public Website", "Home page", "Animated landing with 3D hero, search, featured suppliers & categories", DONE),
    ("Public Website", "About / Contact / FAQs", "Company pages; dynamic FAQ with structured data", DONE),
    ("Public Website", "Supplier directory", "Browse verified suppliers with filters & pagination", DONE),
    ("Public Website", "Supplier details", "Profile, products, trade terms, contact unlock", DONE),
    ("Public Website", "Product catalog", "Advanced filters (category, MOQ, budget) + currency conversion", DONE),
    ("Public Website", "Product details", "Full product page with localized pricing", DONE),
    ("Public Website", "Category system", "Parent/child categories with dynamic landing pages", DONE),
    ("Public Website", "Filter system", "Country, MOQ, budget and verified-only filters", DONE),
    ("Public Website", "Country-based pricing", "Auto currency by visitor country + conversion", DONE),
    ("Public Website", "Legal/content pages", "Privacy, Terms, Cookie (admin-managed CMS)", DONE),
    ("Public Website", "SEO", "Metadata, JSON-LD, sitemap.xml, robots.txt, clean URLs", DONE),
    ("Public Website", "Responsive + dark mode", "Mobile-first responsive UI with light/dark themes", DONE),
    # Auth
    ("Authentication", "Login", "Email/password sign-in", DONE),
    ("Authentication", "Buyer registration", "Quick single-step buyer signup", DONE),
    ("Authentication", "Supplier registration", "Multi-step supplier onboarding wizard", DONE),
    ("Authentication", "Google SSO", "Sign in / up with Google", READY),
    ("Authentication", "Forgot / reset password", "Token-based password reset flow", DONE),
    ("Authentication", "Phone OTP verification", "SMS one-time-code verification", READY),
    ("Authentication", "Roles & route protection", "Buyer / Supplier / Admin sessions; protected areas", DONE),
    # Buyer
    ("Buyer", "Dashboard", "Overview with credits, mode and quick actions", DONE),
    ("Buyer", "Post RFQ", "Create requests for quotation", DONE),
    ("Buyer", "Manage RFQs", "List, edit, close/reopen, delete", DONE),
    ("Buyer", "Manage quotations", "Review and accept/reject supplier quotes", DONE),
    ("Buyer", "Unlock supplier contacts", "Spend credits to reveal contact details", DONE),
    ("Buyer", "Credits & wallet", "Buy credit packs; full transaction history", DONE),
    ("Buyer", "Saved items", "Bookmark suppliers, products and RFQs", DONE),
    ("Buyer", "Notifications", "In-app notifications + channel preferences", DONE),
    ("Buyer", "Profile & password", "Update profile; change password", DONE),
    # Supplier
    ("Supplier", "Dashboard", "Overview with verification status & catalog", DONE),
    ("Supplier", "RFQ feed", "Browse open buyer requests to quote on", DONE),
    ("Supplier", "Submit quotations", "Send competitive quotes; withdraw pending", DONE),
    ("Supplier", "Product management", "Create, edit, delete catalog products", DONE),
    ("Supplier", "Company profile", "Company info, terms, strengths, contacts", DONE),
    ("Supplier", "Verification", "Submit documents; track approval status", DONE),
    ("Supplier", "Profile & password", "Update profile; change password", DONE),
    # Admin
    ("Admin", "Dashboard", "Platform stats + recent activity (audit) feed", DONE),
    ("Admin", "User management", "View, approve, reject, block/unblock users", DONE),
    ("Admin", "Impersonate login", "Securely log in as a buyer/supplier", DONE),
    ("Admin", "Supplier verification", "Review queue; approve/reject + verified badge", DONE),
    ("Admin", "Document review", "Approve/reject individual company documents", DONE),
    ("Admin", "RFQ moderation", "Flag or delete spam requests", DONE),
    ("Admin", "Category management", "Add / edit / delete categories", DONE),
    ("Admin", "CMS management", "Create/edit/delete content & legal pages", DONE),
    ("Admin", "FAQ management", "Create / edit / delete FAQs", DONE),
    ("Admin", "Credit management", "Add/remove credits; review transactions", DONE),
    ("Admin", "Audit logging", "Every admin action recorded", DONE),
    # Payments
    ("Payments & Credits", "Credit packs", "Tiered packs (Starter/Growth/Scale)", DONE),
    ("Payments & Credits", "Stripe Checkout", "Secure card payment for credits", READY),
    ("Payments & Credits", "Payment webhook", "Auto-grant credits on successful payment", READY),
    ("Payments & Credits", "Credit ledger", "Atomic, double-spend-safe transactions", DONE),
    # Notifications
    ("Notifications", "In-app + bell", "Unread badge and notification center", DONE),
    ("Notifications", "Email", "Transactional email (Resend)", READY),
    ("Notifications", "SMS / WhatsApp", "Twilio channels", READY),
    ("Notifications", "Telegram", "Telegram bot notifications", READY),
    ("Notifications", "Channel preferences", "Per-user enable/disable per channel", DONE),
    # Platform
    ("Platform", "Type-safe codebase", "End-to-end TypeScript", DONE),
    ("Platform", "Security", "RBAC, validation, transactions, audit, hashing", DONE),
    ("Platform", "Scalable architecture", "Stateless app + managed data services", DONE),
    ("Platform", "Production build", "Verified compiling production build", DONE),
]

r = 5
for i, (mod, feat, desc, status) in enumerate(features, 1):
    zebra = ZEBRA if i % 2 == 0 else WHITE
    cells = [
        fe.cell(r, 1, i),
        fe.cell(r, 2, mod),
        fe.cell(r, 3, feat),
        fe.cell(r, 4, desc),
    ]
    cells[0].alignment = Alignment(horizontal="center", vertical="center")
    cells[1].font = Font(bold=True, color=TEAL_DARK, size=10)
    cells[1].alignment = Alignment(vertical="center", indent=1)
    cells[2].font = Font(bold=True, size=10, color="222222")
    cells[2].alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    cells[3].font = Font(size=10, color="333333")
    cells[3].alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    for c in cells:
        c.fill = fill(zebra); c.border = border
    st = fe.cell(r, 5, status)
    status_style(st, status)
    fe.row_dimensions[r].height = 26
    r += 1

for col, w in zip("ABCDE", [5, 20, 26, 60, 20]):
    fe.column_dimensions[col].width = w

# Legend
lr = r + 1
fe.cell(lr, 2, "Legend:").font = Font(bold=True, color=GREY)
g = fe.cell(lr, 3, "Completed"); status_style(g, "Completed")
a = fe.cell(lr, 4, "Ready – needs API key (Stripe/Twilio/Resend/Google)")
a.font = Font(color=AMBER, size=10); a.alignment = Alignment(indent=1)

# ============================================================ TECHNOLOGY
te = wb.create_sheet("Technology Stack")
te.sheet_view.showGridLines = False
title_block(te, "Technology Stack", "Tools used and why", 4)
header_row(te, 4, ["#", "Area", "Technology", "Purpose / Why chosen"])
te.freeze_panes = "A5"

tech = [
    ("Framework", "Next.js 16 (App Router, React 19)", "SSR/SSG/ISR for SEO; one full-stack codebase"),
    ("Language", "TypeScript", "Type safety across frontend, backend and database"),
    ("Backend", "Next.js Route Handlers + Server Actions", "Node.js backend — no separate API server to maintain"),
    ("Database", "PostgreSQL", "Relational integrity for RFQs, quotes, credits, transactions"),
    ("ORM", "Prisma 7", "Type-safe database access, schema & migrations"),
    ("Authentication", "Auth.js v5 (NextAuth)", "Email/password, Google SSO, roles, admin impersonation"),
    ("Validation", "Zod", "One schema validates client + server inputs"),
    ("UI / Styling", "Tailwind CSS v4 + shadcn/ui (Radix)", "Custom, accessible design system (not a template)"),
    ("Animation", "GSAP + ScrollTrigger", "Hero timeline & scroll-reveal animations"),
    ("3D Graphics", "Three.js / React Three Fiber", "Interactive 3D hero for a premium feel"),
    ("Smooth Scroll", "Lenis", "Buttery smooth scrolling experience"),
    ("Payments", "Stripe", "Credit-pack checkout + secure webhooks"),
    ("Email", "Resend", "Transactional email notifications"),
    ("SMS / WhatsApp / OTP", "Twilio", "Phone verification & multi-channel alerts"),
    ("Messaging", "Telegram Bot API", "Opt-in Telegram notifications"),
    ("File Storage", "Cloudflare R2 / S3", "Company documents & media"),
    ("Currency / Geo", "Exchange-rate API + IP geolocation", "Country-based pricing & conversion"),
    ("Data Fetching", "React Server Components + TanStack Query", "Fast server data + responsive client caching"),
    ("Hosting", "Any Node.js host (Railway / Render / DigitalOcean / AWS / VPS)", "Standard Next.js app — deploys to any Node host or Docker"),
    ("Database Hosting", "Managed PostgreSQL (Neon / Supabase / RDS / self-hosted)", "Scalable managed Postgres"),
    ("Monitoring", "Sentry + Vercel Analytics", "Error tracking & Web Vitals"),
    ("Tooling", "ESLint, Git", "Code quality and version control"),
]
r = 5
for i, (area, tech_name, why) in enumerate(tech, 1):
    zebra = ZEBRA if i % 2 == 0 else WHITE
    c1 = te.cell(r, 1, i); c2 = te.cell(r, 2, area); c3 = te.cell(r, 3, tech_name); c4 = te.cell(r, 4, why)
    c1.alignment = Alignment(horizontal="center", vertical="center")
    c2.font = Font(bold=True, color=TEAL_DARK, size=10); c2.alignment = Alignment(vertical="center", indent=1)
    c3.font = Font(bold=True, size=10, color="222222"); c3.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    c4.font = Font(size=10, color="333333"); c4.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    for c in (c1, c2, c3, c4):
        c.fill = fill(zebra); c.border = border
    te.row_dimensions[r].height = 24
    r += 1
for col, w in zip("ABCD", [5, 24, 38, 58]):
    te.column_dimensions[col].width = w

# ============================================================ PHASES
ph = wb.create_sheet("Delivery Phases")
ph.sheet_view.showGridLines = False
title_block(ph, "Delivery Phases", "Project broken into delivery milestones", 3)
header_row(ph, 4, ["Phase", "Scope", "Status"])
ph.freeze_panes = "A5"
phases = [
    ("0 — Foundation", "Project setup, database, design system, auth skeleton", DONE),
    ("1 — Public SEO Site", "Listings, detail pages, categories, filters, currency, SEO", DONE),
    ("2 — Authentication", "Login, registration, Google SSO, supplier onboarding", DONE),
    ("3 — Buyer Dashboard", "RFQs, quotations, saved items, credits, profile", DONE),
    ("4 — Supplier Dashboard", "RFQ feed, quotes, products, company, verification", DONE),
    ("5 — Payments & Credits", "Unlock flow, Stripe checkout, credit packs", DONE),
    ("6 — Admin Panel", "Users, verification, CMS, FAQ, categories, credits", DONE),
    ("7 — Notifications & Polish", "Notification center, final polish, documentation", DONE),
]
r = 5
for i, (phase, scope, status) in enumerate(phases, 1):
    zebra = ZEBRA if i % 2 == 0 else WHITE
    c1 = ph.cell(r, 1, phase); c2 = ph.cell(r, 2, scope)
    c1.font = Font(bold=True, color=TEAL_DARK, size=11); c1.alignment = Alignment(vertical="center", indent=1)
    c2.font = Font(size=10, color="333333"); c2.alignment = Alignment(vertical="center", indent=1, wrap_text=True)
    for c in (c1, c2):
        c.fill = fill(zebra); c.border = border
    st = ph.cell(r, 3, status); status_style(st, status)
    ph.row_dimensions[r].height = 26
    r += 1
for col, w in zip("ABC", [26, 64, 16]):
    ph.column_dimensions[col].width = w

out = "BigSeaa_Feature_List_and_Technology.xlsx"
wb.save(out)
print("Saved:", out)
print("Sheets:", wb.sheetnames)
print("Features:", len(features), "| Technologies:", len(tech), "| Phases:", len(phases))
